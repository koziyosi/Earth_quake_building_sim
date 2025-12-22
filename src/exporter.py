"""
Export utilities for simulation results.
"""
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import numpy as np


class ResultsExporter:
    """Export simulation results to various formats."""
    
    @staticmethod
    def export_to_csv(
        filepath: str,
        time_array: np.ndarray,
        data_dict: Dict[str, np.ndarray],
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Export time-series data to CSV.
        
        Args:
            filepath: Output file path
            time_array: Time values
            data_dict: Dictionary of {column_name: data_array}
            metadata: Optional metadata to include as comments
            
        Returns:
            Path to created file
        """
        filepath = Path(filepath)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write metadata as comments
            if metadata:
                for key, value in metadata.items():
                    f.write(f"# {key}: {value}\n")
                f.write(f"# Exported: {datetime.now().isoformat()}\n")
                f.write("#\n")
            
            # Write header
            headers = ['Time(s)'] + list(data_dict.keys())
            writer.writerow(headers)
            
            # Write data
            for i, t in enumerate(time_array):
                row = [f"{t:.4f}"]
                for col_data in data_dict.values():
                    if i < len(col_data):
                        row.append(f"{col_data[i]:.6e}")
                    else:
                        row.append("")
                writer.writerow(row)
                
        return str(filepath)
    
    @staticmethod
    def export_to_json(
        filepath: str,
        results: Dict[str, Any],
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Export results to JSON format.
        
        Args:
            filepath: Output file path
            results: Results dictionary (numpy arrays will be converted to lists)
            metadata: Optional metadata
            
        Returns:
            Path to created file
        """
        filepath = Path(filepath)
        
        # Convert numpy arrays to lists
        def convert(obj):
            if isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, dict):
                return {k: convert(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert(item) for item in obj]
            return obj
        
        output = {
            "metadata": {
                **(metadata or {}),
                "exported": datetime.now().isoformat()
            },
            "results": convert(results)
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2)
            
        return str(filepath)

    @staticmethod
    def export_displacement_history(
        filepath: str,
        time_array: np.ndarray,
        history_u: List[np.ndarray],
        node_ids: Optional[List[int]] = None
    ) -> str:
        """
        Export displacement history.
        
        Args:
            filepath: Output file path
            time_array: Time array
            history_u: List of displacement arrays per time step
            node_ids: Optional list of node IDs to include
            
        Returns:
            Path to created file
        """
        filepath = Path(filepath)
        
        # Flatten history
        if len(history_u) == 0:
            return str(filepath)
            
        n_dof = len(history_u[0])
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            headers = ['Time(s)'] + [f'DOF_{i}' for i in range(n_dof)]
            writer.writerow(headers)
            
            # Data
            for i, t in enumerate(time_array):
                if i < len(history_u):
                    row = [f"{t:.4f}"] + [f"{u:.6e}" for u in history_u[i]]
                    writer.writerow(row)
                    
        return str(filepath)


class OBJExporter:
    """Export 3D model to Wavefront OBJ format."""
    
    @staticmethod
    def export_frame_model(
        filepath: str,
        nodes: List[Any],
        elements: List[Any],
        scale: float = 1.0
    ) -> str:
        """
        Export frame model to OBJ format.
        
        Args:
            filepath: Output file path
            nodes: List of Node objects
            elements: List of Element objects
            scale: Scale factor for coordinates
            
        Returns:
            Path to created file
        """
        filepath = Path(filepath)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"# EarthQuake Building Sim - OBJ Export\n")
            f.write(f"# Exported: {datetime.now().isoformat()}\n\n")
            
            # Node ID to vertex index mapping
            node_to_vertex = {}
            
            # Write vertices
            for i, node in enumerate(nodes):
                v_idx = i + 1  # OBJ uses 1-based indexing
                node_to_vertex[node.id] = v_idx
                f.write(f"v {node.x * scale:.6f} {node.z * scale:.6f} {node.y * scale:.6f}\n")
            
            f.write("\n")
            
            # Write edges as lines
            for elem in elements:
                v1 = node_to_vertex.get(elem.node_i.id)
                v2 = node_to_vertex.get(elem.node_j.id)
                if v1 and v2:
                    f.write(f"l {v1} {v2}\n")
                    
        return str(filepath)


def export_excel(filepath: str, time_array: np.ndarray, 
                 data_dict: Dict[str, np.ndarray]) -> Optional[str]:
    """
    Export to Excel format (requires openpyxl).
    
    Returns None if openpyxl is not available.
    """
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Simulation Results"
        
        # Headers
        headers = ['Time(s)'] + list(data_dict.keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, t in enumerate(time_array, 2):
            ws.cell(row=row, column=1, value=t)
            for col, col_data in enumerate(data_dict.values(), 2):
                if row - 2 < len(col_data):
                    ws.cell(row=row, column=col, value=float(col_data[row - 2]))
        
        wb.save(filepath)
        return filepath
        
    except ImportError:
        print("openpyxl not installed. Excel export unavailable.")
        return None
